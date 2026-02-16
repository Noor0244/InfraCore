# app/routes/daily_entry_ui.py
# --------------------------------------------------
# Daily Activity Execution (UI)
# --------------------------------------------------

from __future__ import annotations

import json
import os
from datetime import date, datetime, timedelta
from pathlib import Path

from fastapi import APIRouter, Request, Depends
from fastapi.responses import RedirectResponse, HTMLResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_, func

from app.utils.template_filters import register_template_filters
from app.utils.dates import parse_date_ddmmyyyy_or_iso

from app.db.session import get_db
from app.models.project import Project
from app.models.project_activity import ProjectActivity
from app.models.project_user import ProjectUser

from app.models.activity import Activity

from app.models.daily_entry import DailyEntry
from app.models.activity_progress import ActivityProgress
from app.models.activity_material import ActivityMaterial
from app.models.material import Material
from app.models.material_stock import MaterialStock
from app.models.material_usage_daily import MaterialUsageDaily

from app.models.daily_work_report import DailyWorkReport
from app.models.daily_work_activity import DailyWorkActivity
from app.models.daily_work_labour import DailyWorkLabour
from app.models.daily_work_machinery import DailyWorkMachinery
from app.models.daily_work_qc import DailyWorkQC
from app.models.daily_work_delay import DailyWorkDelay
from app.models.daily_work_material import DailyWorkMaterial
from app.models.daily_work_upload import DailyWorkUpload
from app.models.road_stretch import RoadStretch

from app.utils.audit_logger import log_action, model_to_dict
from app.utils.activity_units import (
    hours_to_display,
    display_to_hours,
    normalize_display_unit,
    normalize_hours_per_day,
)

router = APIRouter(
    prefix="/execution",
    tags=["Daily Execution UI"]
)

templates = Jinja2Templates(directory="app/templates")
register_template_filters(templates)


# =====================================================
# PROJECT SELECTION
# =====================================================
@router.get("/", response_class=HTMLResponse)
def execution_project_select(
    request: Request,
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=302)

    # Filtering rules:
    # 1) Only active, not completed
    # 2) Exclude known seed/test/preset projects
    # 3) Access control: show projects the user owns OR is assigned to
    base = (
        db.query(Project)
        .outerjoin(ProjectUser, Project.id == ProjectUser.project_id)
        .filter(
            Project.is_active == True,
            Project.status == "active",
            Project.completed_at.is_(None),
            # Exclude known non-real projects (seed/preset/test)
            ~func.lower(Project.name).like("%preset%"),
            ~func.lower(Project.name).like("%test%"),
        )
    )

    if user.get("role") in {"admin", "superadmin"}:
        projects = (
            base.distinct()
            .order_by(Project.created_at.desc())
            .all()
        )
    else:
        projects = (
            base.filter(
                or_(
                    Project.created_by == user["id"],
                    ProjectUser.user_id == user["id"],
                )
            )
            .distinct()
            .order_by(Project.created_at.desc())
            .all()
        )

    today = date.today()
    planned_rows = (
        db.query(ProjectActivity.project_id, func.count(ProjectActivity.id))
        .filter(ProjectActivity.start_date <= today, ProjectActivity.end_date >= today)
        .group_by(ProjectActivity.project_id)
        .all()
    )
    planned_today_counts = {int(pid): int(cnt or 0) for pid, cnt in planned_rows}

    return templates.TemplateResponse(
        "daily_execution_projects.html",
        {
            "request": request,
            "user": user,
            "projects": projects,
            "planned_today_counts": planned_today_counts,
        }
    )


# =====================================================
# DOWNLOAD DPR AS EXCEL
# =====================================================
@router.get("/{project_id}/download-excel")
def download_dpr_excel(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Download Daily Work Execution Report as Excel"""
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=302)

    # Get date from query params
    date_str = request.query_params.get("date", "")
    report_date = date.today()
    if date_str:
        try:
            report_date = parse_date_ddmmyyyy_or_iso(date_str)
        except Exception:
            report_date = date.today()

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        return RedirectResponse("/execution", status_code=302)

    report = (
        db.query(DailyWorkReport)
        .filter(DailyWorkReport.project_id == project_id, DailyWorkReport.report_date == report_date)
        .first()
    )

    # Get activities for the report
    activities = []
    if report:
        activities = (
            db.query(DailyWorkActivity)
            .filter(DailyWorkActivity.report_id == report.id)
            .all()
        )

    # Create Excel file
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "DPR"

        # Header styling
        header_fill = PatternFill(start_color="4F8CFF", end_color="4F8CFF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Daily Work Execution Report - {project.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Project Info
        row = 3
        ws[f'A{row}'] = "Project:"
        ws[f'B{row}'] = project.name
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Date:"
        ws[f'B{row}'] = report_date.strftime("%d/%m/%Y")
        ws[f'A{row}'].font = Font(bold=True)
        
        if report:
            row += 1
            ws[f'A{row}'] = "Weather:"
            ws[f'B{row}'] = report.weather or ""
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Shift:"
            ws[f'B{row}'] = report.shift or ""
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Supervisor:"
            ws[f'B{row}'] = report.supervisor_name or ""
            ws[f'A{row}'].font = Font(bold=True)

        # Activities Table
        row += 2
        headers = ['Activity', 'Planned (Before)', 'Executed Today', 'Cumulative (After)', 'Unit', 'Status']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Activity data
        for activity in activities:
            row += 1
            activity_obj = db.query(Activity).filter(Activity.id == activity.activity_id).first()
            activity_name = activity_obj.name if activity_obj else f"Activity {activity.activity_id}"
            
            ws.cell(row=row, column=1, value=activity_name)
            ws.cell(row=row, column=2, value=activity.planned_before or 0)
            ws.cell(row=row, column=3, value=activity.executed_today or 0)
            ws.cell(row=row, column=4, value=activity.cumulative_after or 0)
            ws.cell(row=row, column=5, value=activity.unit or "")
            ws.cell(row=row, column=6, value=activity.status or "")

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"DPR_{project.name.replace(' ', '_')}_{report_date.strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        return Response(content="Excel library not installed. Please install openpyxl.", status_code=500)
    except Exception as e:
        return Response(content=f"Error generating Excel: {str(e)}", status_code=500)


# =====================================================
# DOWNLOAD DPR AS PDF
# =====================================================
@router.get("/{project_id}/download-pdf")
def download_dpr_pdf(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Download Daily Work Execution Report as PDF"""
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=302)

    # Get date from query params
    date_str = request.query_params.get("date", "")
    report_date = date.today()
    if date_str:
        try:
            report_date = parse_date_ddmmyyyy_or_iso(date_str)
        except Exception:
            report_date = date.today()

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        return RedirectResponse("/execution", status_code=302)

    report = (
        db.query(DailyWorkReport)
        .filter(DailyWorkReport.project_id == project_id, DailyWorkReport.report_date == report_date)
        .first()
    )

    # Get activities for the report
    activities = []
    if report:
        activities = (
            db.query(DailyWorkActivity)
            .filter(DailyWorkActivity.report_id == report.id)
            .all()
        )

    # Create PDF file
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a2332'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Daily Work Execution Report", title_style))
        elements.append(Paragraph(f"{project.name}", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Project Information
        info_data = [
            ["Date:", report_date.strftime("%d/%m/%Y")],
            ["Project ID:", f"#{project.id}"],
        ]
        
        if report:
            info_data.extend([
                ["Weather:", report.weather or "—"],
                ["Shift:", report.shift or "—"],
                ["Supervisor:", report.supervisor_name or "—"],
                ["Work Chainage:", f"{report.work_chainage_from or '—'} → {report.work_chainage_to or '—'}"],
            ])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Activities Table
        if activities:
            elements.append(Paragraph("Activities Executed", styles['Heading3']))
            elements.append(Spacer(1, 0.1*inch))
            
            activity_data = [['Activity', 'Planned\n(Before)', 'Executed\nToday', 'Cumulative\n(After)', 'Unit', 'Status']]
            
            for activity in activities:
                activity_obj = db.query(Activity).filter(Activity.id == activity.activity_id).first()
                activity_name = activity_obj.name if activity_obj else f"Activity {activity.activity_id}"
                
                activity_data.append([
                    activity_name[:40],
                    f"{activity.planned_before or 0:.3f}",
                    f"{activity.executed_today or 0:.3f}",
                    f"{activity.cumulative_after or 0:.3f}",
                    activity.unit or "",
                    activity.status or ""
                ])
            
            activity_table = Table(activity_data, colWidths=[2.5*inch, 0.8*inch, 0.8*inch, 1*inch, 0.6*inch, 0.8*inch])
            activity_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8CFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(activity_table)
        else:
            elements.append(Paragraph("No activities recorded for this date.", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"DPR_{project.name.replace(' ', '_')}_{report_date.strftime('%Y%m%d')}.pdf"
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        return Response(content="PDF library not installed. Please install reportlab.", status_code=500)
    except Exception as e:
        return Response(content=f"Error generating PDF: {str(e)}", status_code=500)


# =====================================================
# DAILY EXECUTION PAGE
# =====================================================
@router.get("/{project_id}", response_class=HTMLResponse)
def daily_execution_page(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=302)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        return RedirectResponse("/execution", status_code=302)

    # Access guard: keep execution scoped to visible projects
    if user.get("role") not in {"admin", "superadmin"}:
        is_owner = (project.created_by == user.get("id"))
        is_member = (
            db.query(ProjectUser)
            .filter(ProjectUser.project_id == project_id, ProjectUser.user_id == user.get("id"))
            .first()
            is not None
        )
        if not (is_owner or is_member):
            return RedirectResponse("/execution", status_code=302)

    # Safety: prevent execution on archived/inactive/completed
    if (not project.is_active) or (project.status != "active") or (project.completed_at is not None):
        return RedirectResponse("/execution", status_code=302)

    # Fetch all available projects for dropdown
    base = (
        db.query(Project)
        .outerjoin(ProjectUser, Project.id == ProjectUser.project_id)
        .filter(
            Project.is_active == True,
        )
    )

    if user.get("role") in {"admin", "superadmin"}:
        all_projects = (
            base.distinct()
            .order_by(Project.name.asc())
            .all()
        )
    else:
        all_projects = (
            base.filter(
                or_(
                    Project.created_by == user["id"],
                    ProjectUser.user_id == user["id"],
                )
            )
            .distinct()
            .order_by(Project.name.asc())
            .all()
        )

    # Report date (querystring): ?date=YYYY-MM-DD
    report_date_str = (request.query_params.get("date") or "").strip()
    report_date = date.today()
    if report_date_str:
        try:
            report_date = parse_date_ddmmyyyy_or_iso(report_date_str)
        except Exception:
            report_date = date.today()

    report = (
        db.query(DailyWorkReport)
        .filter(DailyWorkReport.project_id == project_id, DailyWorkReport.report_date == report_date)
        .first()
    )

    planned = (
        db.query(ProjectActivity)
        .filter(ProjectActivity.project_id == project_id)
        .order_by(ProjectActivity.end_date.asc())
        .all()
    )

    planned_today = []
    planned_today_map: dict[int, float] = {}
    for pa in planned:
        if pa.start_date and pa.end_date and pa.start_date <= report_date <= pa.end_date:
            planned_today.append(pa)
            try:
                duration_days = (pa.end_date - pa.start_date).days + 1
            except Exception:
                duration_days = 1
            if duration_days <= 0:
                duration_days = 1
            planned_today_map[int(pa.activity_id)] = float(pa.planned_quantity or 0) / float(duration_days)

    activity_def_count = (
        db.query(func.count(Activity.id))
        .filter(Activity.project_id == project_id)
        .scalar()
        or 0
    )

    # Stretch context (road-only, safe fallback)
    stretches = (
        db.query(RoadStretch)
        .filter(RoadStretch.project_id == project_id)
        .order_by(RoadStretch.sequence_no.asc())
        .all()
    )

    current_stretch = None
    if stretches:
        for s in stretches:
            start = s.start_date or s.planned_start_date
            end = s.end_date or s.planned_end_date
            if start and end:
                if start <= report_date <= end:
                    current_stretch = s
                    break
            elif start and report_date >= start:
                current_stretch = s
                break
            elif end and report_date <= end:
                current_stretch = s
                break

        if current_stretch is None:
            current_stretch = next((s for s in stretches if s.is_active), stretches[0])

    activity_ids = [pa.activity_id for pa in planned]

    # Cumulative before report date (for validation + display)
    cum_rows = (
        db.query(DailyEntry.activity_id, func.sum(DailyEntry.quantity_done))
        .filter(
            DailyEntry.project_id == project_id,
            DailyEntry.activity_id.in_(activity_ids) if activity_ids else True,
            DailyEntry.entry_date < report_date,
        )
        .group_by(DailyEntry.activity_id)
        .all()
    )
    cum_before = {aid: float(total or 0) for (aid, total) in cum_rows}

    # Prefill executed today (prefer DPR rows, fallback to DailyEntry)
    executed_today = {}
    executed_today_hours = {}
    remarks_today = {}
    approval_status_by_activity = {}
    if report:
        for r in db.query(DailyWorkActivity).filter(DailyWorkActivity.report_id == report.id).all():
            executed_today[int(r.activity_id)] = float(r.executed_today or 0)
            executed_today_hours[int(r.activity_id)] = float(getattr(r, "executed_today_hours", 0.0) or 0.0)
            if r.remarks:
                remarks_today[int(r.activity_id)] = r.remarks
            approval_status_by_activity[int(r.activity_id)] = str(getattr(r, "approval_status", "Changed") or "Changed")
    else:
        day_rows = (
            db.query(DailyEntry)
            .filter(
                DailyEntry.project_id == project_id,
                DailyEntry.entry_date == report_date,
                DailyEntry.activity_id.in_(activity_ids) if activity_ids else True,
            )
            .all()
        )
        for e in day_rows:
            executed_today[int(e.activity_id)] = float(e.quantity_done or 0)
            executed_today_hours[int(e.activity_id)] = float(getattr(e, "quantity_done_hours", 0.0) or 0.0)
            if e.remarks:
                remarks_today[int(e.activity_id)] = e.remarks

    progress_rows = (
        db.query(ActivityProgress)
        .filter(ActivityProgress.project_id == project_id, ActivityProgress.activity_id.in_(activity_ids) if activity_ids else True)
        .all()
    )
    progress_by_activity = {int(ap.activity_id): ap for ap in progress_rows}

    # Time tracking view model for template (no conversion math in Jinja)
    # Keyed by activity_id
    time_by_activity_id: dict[int, dict] = {}

    # Cumulative before report date (time-hours)
    cum_time_rows = (
        db.query(DailyEntry.activity_id, func.sum(DailyEntry.quantity_done_hours))
        .filter(
            DailyEntry.project_id == project_id,
            DailyEntry.activity_id.in_(activity_ids) if activity_ids else True,
            DailyEntry.entry_date < report_date,
        )
        .group_by(DailyEntry.activity_id)
        .all()
    )
    cum_before_hours = {int(aid): float(total or 0) for (aid, total) in cum_time_rows}

    # Precompute per activity from Activity table (joined via ProjectActivity.activity)
    for pa in planned:
        act = getattr(pa, "activity", None)
        if not act:
            continue
        aid = int(pa.activity_id)
        du = normalize_display_unit(getattr(act, "display_unit", None))
        hpd = normalize_hours_per_day(getattr(act, "hours_per_day", None), default=8.0)
        planned_hours = float(getattr(act, "planned_quantity_hours", 0.0) or 0.0)
        cbh = float(cum_before_hours.get(aid, 0.0) or 0.0)
        eth = float(executed_today_hours.get(aid, 0.0) or 0.0)
        ca_h = cbh + eth
        bal_h = max(planned_hours - ca_h, 0.0)

        time_by_activity_id[aid] = {
            "planned_hours": planned_hours,
            "cum_before_hours": cbh,
            "executed_today_hours": eth,
            "display_unit": du,
            "hours_per_day": hpd,
            "planned_display": round(hours_to_display(planned_hours, du, hpd), 3),
            "executed_today_display": round(hours_to_display(eth, du, hpd), 3),
            "cum_after_display": round(hours_to_display(ca_h, du, hpd), 3),
            "balance_display": round(hours_to_display(bal_h, du, hpd), 3),
        }

    # Activity→Material mapping
    am_rows = (
        db.query(ActivityMaterial)
        .filter(ActivityMaterial.activity_id.in_(activity_ids) if activity_ids else True)
        .all()
    )

    activity_material_map: dict[str, list[dict]] = {}
    material_ids: set[int] = set()
    for am in am_rows:
        activity_material_map.setdefault(str(am.activity_id), []).append(
            {
                "material_id": int(am.material_id),
                "rate": float(am.consumption_rate or 0),
            }
        )
        material_ids.add(int(am.material_id))

    materials = []
    material_stock = {}
    if material_ids:
        materials = db.query(Material).filter(Material.id.in_(sorted(material_ids))).order_by(Material.name.asc()).all()
        stock_rows = (
            db.query(MaterialStock)
            .filter(MaterialStock.project_id == project_id, MaterialStock.material_id.in_(sorted(material_ids)))
            .all()
        )
        material_stock = {int(s.material_id): float(s.quantity_available or 0) for s in stock_rows}

    materials_json = json.dumps([
        {
            "id": int(m.id),
            "name": str(m.name or ""),
            "unit": str(m.unit or ""),
        }
        for m in materials
    ])

    existing_materials = {}
    if report:
        for mr in report.material_rows:
            existing_materials[int(mr.material_id)] = mr

    return templates.TemplateResponse(
        "daily_entry.html",
        {
            "request": request,
            "user": user,
            "project": project,
            "all_projects": all_projects,
            "report": report,
            "report_date": report_date,
            "prev_date": report_date - timedelta(days=1),
            "next_date": report_date + timedelta(days=1),
            "planned": planned,
            "planned_today": planned_today,
            "planned_today_json": json.dumps(planned_today_map),
            "activity_def_count": int(activity_def_count),
            "current_stretch": current_stretch,
            "cum_before": cum_before,
            "executed_today": executed_today,
            "remarks_today": remarks_today,
            "approval_status_by_activity": approval_status_by_activity,
            "progress_by_activity": progress_by_activity,
            "activity_material_map_json": json.dumps(activity_material_map),
            "materials_json": materials_json,
            "materials": materials,
            "material_stock": material_stock,
            "existing_materials": existing_materials,
            "time_by_activity_id": time_by_activity_id,
            "today": date.today(),
        },
    )


@router.post("/{project_id}")
async def save_daily_work_report(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=302)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        return RedirectResponse("/execution", status_code=302)

    if user.get("role") != "admin":
        is_owner = (project.created_by == user.get("id"))
        is_member = (
            db.query(ProjectUser)
            .filter(ProjectUser.project_id == project_id, ProjectUser.user_id == user.get("id"))
            .first()
            is not None
        )
        if not (is_owner or is_member):
            return RedirectResponse("/execution", status_code=302)

    form = await request.form()

    # ---------------- Header ----------------
    report_date_raw = (form.get("report_date") or "").strip()
    try:
        report_date = parse_date_ddmmyyyy_or_iso(report_date_raw)
    except Exception:
        return RedirectResponse(f"/execution/{project_id}", status_code=302)

    def _s(key: str, max_len: int | None = None) -> str:
        v = (form.get(key) or "")
        v = str(v).strip()
        if max_len:
            v = v[:max_len]
        return v

    weather = _s("weather", 20) or "Sunny"
    shift = _s("shift", 20) or "Day"
    work_chainage_from = _s("work_chainage_from", 50)
    work_chainage_to = _s("work_chainage_to", 50)
    supervisor_name = _s("supervisor_name", 150)
    engineer_remarks = _s("engineer_remarks", 2000)

    toolbox_talk_conducted = (form.get("toolbox_talk_conducted") == "on")
    incident_or_near_miss = (form.get("incident_or_near_miss") == "on")
    incident_description = _s("incident_description", 2000)
    ppe_raw = _s("ppe_compliance_percent", 20)
    ppe_compliance_percent = None
    if ppe_raw:
        try:
            ppe_compliance_percent = float(ppe_raw)
        except Exception:
            ppe_compliance_percent = None

    action = (_s("action", 20) or "save").lower()
    if action == "approve":
        status = "Approved"
    elif action == "submit":
        status = "Submitted"
    else:
        status = "Draft"

    existing = (
        db.query(DailyWorkReport)
        .filter(DailyWorkReport.project_id == project_id, DailyWorkReport.report_date == report_date)
        .first()
    )
    is_create = existing is None

    old_snapshot = None
    if existing:
        old_snapshot = {
            "report": model_to_dict(existing),
            "activities": [model_to_dict(r) for r in db.query(DailyWorkActivity).filter(DailyWorkActivity.report_id == existing.id).all()],
            "materials": [model_to_dict(r) for r in existing.material_rows],
        }

        # Reverse previous material consumption from stock (idempotent saves)
        for mr in list(existing.material_rows):
            prev_consumed = float(mr.consumed_today or 0)
            if prev_consumed > 0:
                stock = (
                    db.query(MaterialStock)
                    .filter(MaterialStock.project_id == project_id, MaterialStock.material_id == mr.material_id)
                    .first()
                )
                if stock:
                    stock.quantity_available = float(stock.quantity_available or 0) + prev_consumed
                    stock.last_updated = datetime.utcnow()

                mud = (
                    db.query(MaterialUsageDaily)
                    .filter(
                        MaterialUsageDaily.project_id == project_id,
                        MaterialUsageDaily.material_id == mr.material_id,
                        MaterialUsageDaily.usage_date == report_date,
                    )
                    .first()
                )
                if mud:
                    mud.quantity_used = 0.0

        # Clear child rows
        existing.labour_rows = []
        existing.machinery_rows = []
        existing.qc_rows = []
        existing.delay_rows = []
        existing.material_rows = []
        existing.uploads = []
        db.query(DailyWorkActivity).filter(DailyWorkActivity.report_id == existing.id).delete()

        report = existing
    else:
        report = DailyWorkReport(project_id=project_id, report_date=report_date)
        db.add(report)

    # Update header
    report.weather = weather
    report.shift = shift
    report.work_chainage_from = work_chainage_from
    report.work_chainage_to = work_chainage_to
    report.supervisor_name = supervisor_name
    report.engineer_remarks = engineer_remarks

    report.toolbox_talk_conducted = toolbox_talk_conducted
    report.ppe_compliance_percent = ppe_compliance_percent
    report.incident_or_near_miss = incident_or_near_miss
    report.incident_description = incident_description

    report.status = status
    report.prepared_by_user_id = user.get("id")
    report.prepared_by_name = user.get("username")
    if status == "Approved":
        report.checked_by = user.get("username")
        report.approved_by = user.get("username")

    db.flush()  # ensure report.id

    # ---------------- Activities ----------------
    activity_id_values = [int(x) for x in (form.getlist("activity_ids") or []) if str(x).strip().isdigit()]
    planned_q = db.query(ProjectActivity).filter(ProjectActivity.project_id == project_id)
    if activity_id_values:
        planned_q = planned_q.filter(ProjectActivity.activity_id.in_(activity_id_values))
    planned_rows = planned_q.all()
    planned_by_activity = {int(pa.activity_id): pa for pa in planned_rows}

    # Time config per activity
    act_q = db.query(Activity).filter(Activity.project_id == project_id)
    if activity_id_values:
        act_q = act_q.filter(Activity.id.in_(activity_id_values))
    act_rows = act_q.all()
    act_by_id = {int(a.id): a for a in act_rows}

    # Cumulative before report date for validation
    cum_q = (
        db.query(DailyEntry.activity_id, func.sum(DailyEntry.quantity_done))
        .filter(DailyEntry.project_id == project_id, DailyEntry.entry_date < report_date)
    )
    if activity_id_values:
        cum_q = cum_q.filter(DailyEntry.activity_id.in_(activity_id_values))
    cum_rows = cum_q.group_by(DailyEntry.activity_id).all()
    cum_before = {int(aid): float(total or 0) for (aid, total) in cum_rows}

    cum_hours_q = (
        db.query(DailyEntry.activity_id, func.sum(DailyEntry.quantity_done_hours))
        .filter(DailyEntry.project_id == project_id, DailyEntry.entry_date < report_date)
    )
    if activity_id_values:
        cum_hours_q = cum_hours_q.filter(DailyEntry.activity_id.in_(activity_id_values))
    cum_hours_rows = cum_hours_q.group_by(DailyEntry.activity_id).all()
    cum_before_hours = {int(aid): float(total or 0) for (aid, total) in cum_hours_rows}

    progress_q = db.query(ActivityProgress).filter(ActivityProgress.project_id == project_id)
    if activity_id_values:
        progress_q = progress_q.filter(ActivityProgress.activity_id.in_(activity_id_values))
    progress_rows = progress_q.all()
    progress_by_activity = {int(ap.activity_id): ap for ap in progress_rows}

    touched_activity_ids: set[int] = set()
    pending_audit_events: list[dict] = []

    for aid in activity_id_values:
        pa = planned_by_activity.get(aid)
        if not pa:
            continue

        act = act_by_id.get(aid)
        act_display_unit = normalize_display_unit(getattr(act, "display_unit", None) if act else None)
        act_hpd = normalize_hours_per_day(getattr(act, "hours_per_day", None) if act else None, default=8.0)
        planned_hours = float(getattr(act, "planned_quantity_hours", 0.0) if act else 0.0) or 0.0

        exec_raw = str(form.get(f"act_exec_{aid}") or "0").strip()
        try:
            executed = float(exec_raw or 0)
        except Exception:
            executed = 0.0

        if executed < 0:
            executed = 0.0

        # ---- Time execution (days/hours input -> hours base) ----
        # Unit selector is per activity row, but must never change stored base values.
        row_unit = normalize_display_unit(str(form.get(f"act_time_unit_{aid}") or act_display_unit))
        exec_time_display_raw = str(form.get(f"act_exec_time_{aid}") or "0").strip()
        try:
            exec_time_display = float(exec_time_display_raw or 0)
        except Exception:
            exec_time_display = 0.0

        # Prefer hidden base-hours (kept in sync by JS). Fallback: convert from display.
        exec_hours_raw = str(form.get(f"act_exec_hours_{aid}") or "0").strip()
        try:
            exec_hours = float(exec_hours_raw or 0)
        except Exception:
            exec_hours = 0.0

        if exec_hours <= 0 and exec_time_display > 0:
            exec_hours = float(display_to_hours(exec_time_display, row_unit, act_hpd) or 0)

        if exec_hours < 0:
            exec_hours = 0.0

        remarks = _s(f"act_rem_{aid}", 2000)
        planned_qty = float(pa.planned_quantity or 0)
        remaining = max(planned_qty - float(cum_before.get(aid, 0)), 0)
        if executed > remaining + 1e-9:
            executed = remaining

        # Time validation: executed hours cannot exceed remaining hours
        remaining_hours = max(planned_hours - float(cum_before_hours.get(aid, 0)), 0.0)
        if exec_hours > remaining_hours + 1e-9:
            exec_hours = remaining_hours

        # Persist unit change; defer audit log until after commit
        if act and str(row_unit) != str(act_display_unit):
            old_unit = act_display_unit
            act.display_unit = row_unit
            pending_audit_events.append(
                {
                    "kind": "activity_unit_change",
                    "activity_id": int(act.id),
                    "old_unit": old_unit,
                    "new_unit": row_unit,
                }
            )

        ap = progress_by_activity.get(aid)
        planned_end = None
        planned_start = None
        if ap:
            planned_start = ap.planned_start
            planned_end = ap.planned_end
        else:
            planned_start = pa.start_date
            planned_end = pa.end_date

        approval_status = _s(f"act_status_{aid}", 20) or "Changed"

        report_row = DailyWorkActivity(
            report_id=report.id,
            project_id=project_id,
            activity_id=aid,
            planned_quantity=planned_qty,
            unit=pa.unit,
            planned_start=(planned_start.isoformat() if planned_start else None),
            planned_end=(planned_end.isoformat() if planned_end else None),
            executed_today=executed,
            planned_quantity_hours=float(planned_hours or 0.0),
            display_unit=str(row_unit),
            hours_per_day=float(act_hpd),
            executed_today_hours=float(exec_hours or 0.0),
            remarks=remarks or None,
            approval_status=approval_status,
        )
        db.add(report_row)

        # Keep legacy progress logic in sync: upsert DailyEntry for this date
        existing_entry = (
            db.query(DailyEntry)
            .filter(
                DailyEntry.project_id == project_id,
                DailyEntry.activity_id == aid,
                DailyEntry.entry_date == report_date,
            )
            .first()
        )

        prev_entry_qty = float(getattr(existing_entry, "quantity_done", 0.0) or 0.0) if existing_entry else 0.0
        prev_entry_hours = float(getattr(existing_entry, "quantity_done_hours", 0.0) or 0.0) if existing_entry else 0.0

        if executed > 0 or exec_hours > 0:
            if existing_entry:
                existing_entry.quantity_done = executed
                existing_entry.quantity_done_hours = float(exec_hours or 0.0)
                existing_entry.remarks = remarks or None
            else:
                db.add(
                    DailyEntry(
                        project_id=project_id,
                        activity_id=aid,
                        quantity_done=executed,
                        quantity_done_hours=float(exec_hours or 0.0),
                        remarks=remarks or None,
                        entry_date=report_date,
                    )
                )

            if (abs(prev_entry_hours - float(exec_hours or 0.0)) > 1e-9) or (abs(prev_entry_qty - float(executed or 0.0)) > 1e-9):
                pending_audit_events.append(
                    {
                        "kind": "activity_execution_time",
                        "activity_id": int(aid),
                        "date": report_date.isoformat(),
                        "old_hours": prev_entry_hours,
                        "new_hours": float(exec_hours or 0.0),
                        "row_unit": row_unit,
                        "hours_per_day": float(act_hpd),
                    }
                )
        else:
            # No execution today: remove legacy entry if present
            if existing_entry:
                db.delete(existing_entry)

                if prev_entry_hours > 0 or prev_entry_qty > 0:
                    pending_audit_events.append(
                        {
                            "kind": "activity_execution_time",
                            "activity_id": int(aid),
                            "date": report_date.isoformat(),
                            "old_hours": prev_entry_hours,
                            "new_hours": 0.0,
                            "row_unit": row_unit,
                            "hours_per_day": float(act_hpd),
                        }
                    )

        touched_activity_ids.add(aid)

    db.flush()

    # Recompute ActivityProgress for touched activities
    for aid in touched_activity_ids:
        pa = planned_by_activity.get(aid)
        ap = progress_by_activity.get(aid)
        if not pa or not ap:
            continue

        cumulative_qty = (
            db.query(func.sum(DailyEntry.quantity_done))
            .filter(DailyEntry.project_id == project_id, DailyEntry.activity_id == aid)
            .scalar()
            or 0.0
        )

        cumulative_hours = (
            db.query(func.sum(DailyEntry.quantity_done_hours))
            .filter(DailyEntry.project_id == project_id, DailyEntry.activity_id == aid)
            .scalar()
            or 0.0
        )

        act = act_by_id.get(aid)
        planned_hours = float(getattr(act, "planned_quantity_hours", 0.0) if act else 0.0) or 0.0

        planned_qty = float(pa.planned_quantity or 0)
        # Prefer time-based progress when a time plan exists; otherwise fallback to legacy quantity-based.
        if planned_hours > 0:
            progress_pct = int(min((cumulative_hours / planned_hours) * 100, 100)) if planned_hours > 0 else 0
        else:
            progress_pct = int(min((cumulative_qty / planned_qty) * 100, 100)) if planned_qty > 0 else 0
        ap.progress_percent = progress_pct

        if progress_pct > 0 and ap.actual_start is None:
            ap.actual_start = report_date

        if progress_pct >= 100:
            ap.actual_end = report_date
            ap.status = "COMPLETED"
        else:
            if progress_pct == 0:
                ap.status = "NOT_STARTED"
            else:
                ap.status = "DELAYED" if report_date > ap.planned_end else "IN_PROGRESS"

        # Keep Activity.executed_quantity_hours in sync (cumulative)
        if act:
            act.executed_quantity_hours = float(cumulative_hours or 0.0)

    # ---------------- Materials ----------------
    material_id_values = [int(x) for x in (form.getlist("material_ids") or []) if str(x).strip().isdigit()]

    for mid in material_id_values:
        planned_today_raw = str(form.get(f"mat_planned_{mid}") or "0").strip()
        issued_raw = str(form.get(f"mat_issued_{mid}") or "0").strip()
        consumed_raw = str(form.get(f"mat_consumed_{mid}") or "0").strip()

        try:
            planned_today = float(planned_today_raw or 0)
        except Exception:
            planned_today = 0.0
        try:
            issued_today = float(issued_raw or 0)
        except Exception:
            issued_today = 0.0
        try:
            consumed_today = float(consumed_raw or 0)
        except Exception:
            consumed_today = 0.0

        if planned_today < 0:
            planned_today = 0.0
        if issued_today < 0:
            issued_today = 0.0
        if consumed_today < 0:
            consumed_today = 0.0

        unit = _s(f"mat_unit_{mid}", 50)
        source = _s(f"mat_source_{mid}", 50)

        if planned_today <= 0 and issued_today <= 0 and consumed_today <= 0:
            continue

        stock = (
            db.query(MaterialStock)
            .filter(MaterialStock.project_id == project_id, MaterialStock.material_id == mid)
            .first()
        )
        available = float(stock.quantity_available or 0) if stock else 0.0
        if consumed_today > available + 1e-9:
            consumed_today = available

        if stock and consumed_today > 0:
            stock.quantity_available = max(float(stock.quantity_available or 0) - consumed_today, 0)
            stock.last_updated = datetime.utcnow()

        report.material_rows.append(
            DailyWorkMaterial(
                report_id=report.id,
                project_id=project_id,
                material_id=mid,
                unit=unit,
                planned_today=planned_today,
                issued_today=issued_today,
                consumed_today=consumed_today,
                source=source or None,
            )
        )

        # Keep daily usage table in sync (upsert)
        if consumed_today > 0:
            mud = (
                db.query(MaterialUsageDaily)
                .filter(
                    MaterialUsageDaily.project_id == project_id,
                    MaterialUsageDaily.material_id == mid,
                    MaterialUsageDaily.usage_date == report_date,
                )
                .first()
            )
            if mud:
                mud.quantity_used = consumed_today
            else:
                db.add(
                    MaterialUsageDaily(
                        project_id=project_id,
                        material_id=mid,
                        usage_date=report_date,
                        quantity_used=consumed_today,
                    )
                )

    # ---------------- Dynamic JSON sections ----------------
    def _json_list(key: str) -> list[dict]:
        raw = form.get(key) or "[]"
        try:
            value = json.loads(raw)
            return value if isinstance(value, list) else []
        except Exception:
            return []

    for row in _json_list("labour_json"):
        report.labour_rows.append(
            DailyWorkLabour(
                report_id=report.id,
                category=str(row.get("category") or "")[:50],
                workers=int(float(row.get("workers") or 0)),
                hours=float(row.get("hours") or 0),
                overtime_hours=float(row.get("overtime_hours") or 0),
                agency=str(row.get("agency") or "")[:150] or None,
            )
        )

    for row in _json_list("machinery_json"):
        report.machinery_rows.append(
            DailyWorkMachinery(
                report_id=report.id,
                equipment_name=str(row.get("equipment_name") or "")[:150],
                hours_used=float(row.get("hours_used") or 0),
                idle_hours=float(row.get("idle_hours") or 0),
                fuel_consumed=float(row.get("fuel_consumed") or 0),
                breakdown=bool(row.get("breakdown")),
                breakdown_remarks=str(row.get("breakdown_remarks") or "")[:2000] or None,
            )
        )

    for row in _json_list("qc_json"):
        report.qc_rows.append(
            DailyWorkQC(
                report_id=report.id,
                test_type=str(row.get("test_type") or "")[:100],
                location=str(row.get("location") or "")[:150] or None,
                test_value=str(row.get("test_value") or "")[:150] or None,
                result=str(row.get("result") or "")[:20] or None,
                engineer_name=str(row.get("engineer_name") or "")[:150] or None,
            )
        )

    for row in _json_list("delays_json"):
        report.delay_rows.append(
            DailyWorkDelay(
                report_id=report.id,
                delay_type=str(row.get("delay_type") or "")[:50],
                start_time=str(row.get("start_time") or "")[:10] or None,
                end_time=str(row.get("end_time") or "")[:10] or None,
                responsible_party=str(row.get("responsible_party") or "")[:150] or None,
                impact_hours=float(row.get("impact_hours") or 0),
                impact_quantity=float(row.get("impact_quantity") or 0),
                remarks=str(row.get("remarks") or "")[:2000] or None,
            )
        )

    # ---------------- Uploads ----------------
    upload_category = _s("upload_category", 50) or "Docs"
    upload_dir = Path("data") / "uploads" / "dpr" / str(project_id) / report_date.isoformat()
    try:
        upload_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    for up in (form.getlist("uploads") or []):
        try:
            filename = getattr(up, "filename", None) or "upload"
            safe_name = filename.replace("..", "_").replace("/", "_").replace("\\", "_")
            stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
            target = upload_dir / f"{stamp}_{safe_name}"
            content = await up.read()
            with open(target, "wb") as f:
                f.write(content)

            rel_path = str(target.as_posix())
            report.uploads.append(
                DailyWorkUpload(
                    report_id=report.id,
                    category=upload_category,
                    file_path=rel_path,
                    original_name=filename,
                )
            )
        except Exception:
            continue

    db.commit()

    log_action(
        db=db,
        request=request,
        action="CREATE" if is_create else "UPDATE",
        entity_type="DailyWorkReport",
        entity_id=report.id,
        description=f"DPR saved: project #{project_id} date {report_date.isoformat()} status {report.status}",
        old_value=old_snapshot,
        new_value={
            "report": model_to_dict(report),
            "activity_count": len(activity_id_values),
            "material_count": len(report.material_rows),
            "status": report.status,
        },
    )

    # Deferred audit events (avoid mid-transaction commits)
    for ev in pending_audit_events:
        try:
            if ev.get("kind") == "activity_unit_change":
                log_action(
                    db=db,
                    request=request,
                    action="UPDATE",
                    entity_type="Activity",
                    entity_id=int(ev.get("activity_id") or 0) or None,
                    description=f"Activity display unit changed (execution): activity #{ev.get('activity_id')} {ev.get('old_unit')} -> {ev.get('new_unit')}",
                    old_value={"activity_id": ev.get("activity_id"), "old_unit": ev.get("old_unit")},
                    new_value={"activity_id": ev.get("activity_id"), "new_unit": ev.get("new_unit")},
                )
            elif ev.get("kind") == "activity_execution_time":
                log_action(
                    db=db,
                    request=request,
                    action="UPDATE",
                    entity_type="ActivityExecutionTime",
                    entity_id=int(ev.get("activity_id") or 0) or None,
                    description=f"Time execution recorded: activity #{ev.get('activity_id')} date {ev.get('date')}",
                    old_value={"activity_id": ev.get("activity_id"), "date": ev.get("date"), "hours": ev.get("old_hours")},
                    new_value={
                        "activity_id": ev.get("activity_id"),
                        "date": ev.get("date"),
                        "hours": ev.get("new_hours"),
                        "row_unit": ev.get("row_unit"),
                        "hours_per_day": ev.get("hours_per_day"),
                    },
                )
        except Exception:
            continue

    return RedirectResponse(f"/execution/{project_id}?date={report_date.isoformat()}", status_code=302)
