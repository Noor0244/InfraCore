from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.dependencies import get_current_user, require_roles
from app.db.session import SessionLocal
from app.models.location import Location
from app.models.project import Project
from app.models.road_stretch import RoadStretch
from app.models.road_geometry import RoadGeometry
from app.models.pavement_design import PavementDesign
from app.schemas.road_schema import (
    DprDataOut,
    PavementDesignIn,
    PavementDesignOut,
    RoadGeometryIn,
    RoadGeometryOut,
    RoadProjectCreate,
    RoadProjectOut,
    RoadStretchCreate,
    RoadStretchOut,
    RoadStretchUpdate,
)
from app.services.road_dpr_service import build_dpr_dataset

router = APIRouter(prefix="/api/road", tags=["Road Projects"])


# ---------------- DB ----------------

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------------- HELPERS ----------------

def _apply_defaults(payload: dict[str, Any], defaults: dict[str, Any] | None, enabled: bool) -> dict[str, Any]:
    if not enabled or not defaults:
        return payload

    merged = dict(defaults)
    merged.update({k: v for k, v in payload.items() if v is not None})
    return merged


def _validate_pavement_design(design: PavementDesignIn, engineering_type: str) -> None:
    et = engineering_type.strip() if engineering_type else ""

    effective_type = et
    if et in {"Urban", "Rural"}:
        if not design.pavement_type:
            raise HTTPException(status_code=422, detail="Pavement type is required for Urban/Rural engineering context.")
        effective_type = design.pavement_type.value

    flexible_fields = [design.gsb_thickness_mm, design.wmm_thickness_mm, design.dbm_thickness_mm, design.bc_thickness_mm]
    rigid_fields = [
        design.slab_thickness_mm,
        design.concrete_grade,
        design.joint_spacing_m,
        design.dowel_diameter_mm,
        design.tie_bar_diameter_mm,
        design.k_value,
    ]
    overlay_fields = [design.existing_pavement_type, design.overlay_thickness_mm]

    if effective_type == "Flexible":
        if any(v is not None for v in rigid_fields):
            raise HTTPException(status_code=422, detail="Rigid-only fields must be null for Flexible pavement.")
        if any(v is None for v in flexible_fields):
            raise HTTPException(status_code=422, detail="Flexible layer thicknesses are required.")

    if effective_type == "Rigid":
        if any(v is not None for v in flexible_fields):
            raise HTTPException(status_code=422, detail="Flexible-only fields must be null for Rigid pavement.")
        required = [design.slab_thickness_mm, design.dowel_diameter_mm, design.tie_bar_diameter_mm]
        if any(v is None for v in required):
            raise HTTPException(status_code=422, detail="Rigid slab and dowel/tie bar fields are required.")

    if effective_type == "Overlay":
        if any(v is None for v in overlay_fields):
            raise HTTPException(status_code=422, detail="Overlay requires existing pavement type and thickness.")


# ---------------- API ----------------


@router.post("/projects", response_model=RoadProjectOut)
def create_road_project(
    payload: RoadProjectCreate,
    user: dict = Depends(require_roles(["admin", "engineer"])),
    db: Session = Depends(get_db),
):
    location = Location(
        country=payload.location.country,
        state=payload.location.state,
        district=payload.location.district,
        city=payload.location.city,
    )
    db.add(location)
    db.flush()

    project = Project(
        name=payload.project_name,
        project_type="Road",
        created_by=int(user["id"]),
        road_type="Road",
        lanes=0,
        road_width=0.0,
        road_length_km=0.0,
        country=payload.location.country,
        state=payload.location.state,
        district=payload.location.district,
        city=payload.location.city,
        planned_start_date=datetime.utcnow().date(),
        planned_end_date=datetime.utcnow().date(),
        location_id=location.id,
    )
    db.add(project)
    db.commit()
    db.refresh(project)

    return RoadProjectOut(
        id=project.id,
        project_name=project.name,
        project_type=project.project_type,
        created_by=project.created_by,
        created_at=project.created_at,
        updated_at=project.updated_at,
        location=location,
    )


@router.post("/projects/{project_id}/stretches", response_model=RoadStretchOut)
def add_road_stretch(
    project_id: int,
    payload: RoadStretchCreate,
    user: dict = Depends(require_roles(["admin", "engineer"])),
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    next_seq = (
        db.query(func.max(RoadStretch.sequence_no))
        .filter(RoadStretch.project_id == project_id)
        .scalar()
        or 0
    )
    sequence_no = int(next_seq) + 1
    stretch = RoadStretch(
        project_id=project_id,
        stretch_code=f"S{sequence_no:02d}",
        stretch_name=payload.stretch_name,
        road_category=payload.road_category.value,
        engineering_type=payload.engineering_type.value,
        start_chainage=payload.start_chainage,
        end_chainage=payload.end_chainage,
        length_m=payload.length_m,
        sequence_no=sequence_no,
        start_date=payload.start_date,
        end_date=payload.end_date,
        is_active=True,
    )
    db.add(stretch)
    db.commit()
    db.refresh(stretch)

    return RoadStretchOut.from_orm(stretch)


@router.put("/stretches/{stretch_id}", response_model=RoadStretchOut)
def update_road_stretch(
    stretch_id: int,
    payload: RoadStretchUpdate,
    user: dict = Depends(require_roles(["admin", "engineer"])),
    db: Session = Depends(get_db),
):
    stretch = db.query(RoadStretch).filter(RoadStretch.id == stretch_id).first()
    if not stretch:
        raise HTTPException(status_code=404, detail="Stretch not found")
    if not (stretch.engineering_type or "").strip():
        raise HTTPException(status_code=422, detail="Engineering type is required for pavement design.")

    for field, value in payload.dict(exclude_unset=True).items():
        if field in {"road_category", "engineering_type"} and value is not None:
            setattr(stretch, field, value.value)
        else:
            setattr(stretch, field, value)

    db.commit()
    db.refresh(stretch)

    return RoadStretchOut.from_orm(stretch)


@router.put("/stretches/{stretch_id}/geometry", response_model=RoadGeometryOut)
def save_geometry(
    stretch_id: int,
    payload: RoadGeometryIn,
    user: dict = Depends(require_roles(["admin", "engineer"])),
    db: Session = Depends(get_db),
):
    stretch = db.query(RoadStretch).filter(RoadStretch.id == stretch_id).first()
    if not stretch:
        raise HTTPException(status_code=404, detail="Stretch not found")

    geometry = db.query(RoadGeometry).filter(RoadGeometry.stretch_id == stretch_id).first()
    if not geometry:
        geometry = RoadGeometry(stretch_id=stretch_id)
        db.add(geometry)

    for field, value in payload.dict(exclude_unset=True).items():
        setattr(geometry, field, value)

    db.commit()
    db.refresh(geometry)

    return RoadGeometryOut.from_orm(geometry)


@router.put("/stretches/{stretch_id}/pavement-design", response_model=PavementDesignOut)
def save_pavement_design(
    stretch_id: int,
    payload: PavementDesignIn,
    user: dict = Depends(require_roles(["admin", "engineer"])),
    db: Session = Depends(get_db),
):
    stretch = db.query(RoadStretch).filter(RoadStretch.id == stretch_id).first()
    if not stretch:
        raise HTTPException(status_code=404, detail="Stretch not found")

    effective_payload = _apply_defaults(payload.dict(exclude_unset=True), payload.defaults, payload.apply_defaults)
    effective_payload.pop("apply_defaults", None)
    effective_payload.pop("defaults", None)

    _validate_pavement_design(PavementDesignIn(**effective_payload), stretch.engineering_type or "")

    pavement = db.query(PavementDesign).filter(PavementDesign.stretch_id == stretch_id).first()
    if not pavement:
        pavement = PavementDesign(stretch_id=stretch_id)
        db.add(pavement)

    for field, value in effective_payload.items():
        setattr(pavement, field, value.value if hasattr(value, "value") else value)

    db.commit()
    db.refresh(pavement)

    return PavementDesignOut.from_orm(pavement)


@router.get("/projects/{project_id}/dpr", response_model=DprDataOut)
def get_dpr_data(
    project_id: int,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    dataset = build_dpr_dataset(db, project_id)
    project = dataset["project"]

    location = None
    if getattr(project, "location_id", None):
        location = db.query(Location).filter(Location.id == project.location_id).first()
    if location:
        location_payload = location
    else:
        location_payload = {
            "id": 0,
            "country": project.country,
            "state": project.state,
            "district": project.district,
            "city": project.city,
        }

    return {
        "project": {
            "id": project.id,
            "project_name": project.name,
            "project_type": project.project_type,
            "created_by": project.created_by,
            "created_at": project.created_at,
            "updated_at": project.updated_at,
            "location": location_payload,
        },
        "stretches": dataset["stretches"],
        "chainage_summary": dataset["chainage_summary"],
        "pavement_tables": dataset["pavement_tables"],
    }


@router.get("/projects/{project_id}/export/excel")
def export_dpr_excel(
    project_id: int,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook

    dataset = build_dpr_dataset(db, project_id)
    project = dataset["project"]

    wb = Workbook()
    ws_project = wb.active
    ws_project.title = "Project"
    ws_project.append(["Project Name", project.name])
    ws_project.append(["Project Type", project.project_type])
    ws_project.append(["Created By", project.created_by])

    ws_stretches = wb.create_sheet("Stretches")
    ws_stretches.append(
        ["Stretch", "Category", "Engineering Type", "Start", "End", "Length (m)"]
    )
    for item in dataset["stretches"]:
        stretch = item["stretch"]
        ws_stretches.append(
            [
                stretch.stretch_name,
                stretch.road_category,
                stretch.engineering_type,
                stretch.start_chainage,
                stretch.end_chainage,
                stretch.length_m,
            ]
        )

    ws_pavement = wb.create_sheet("Pavement")
    ws_pavement.append(["Stretch", "Engineering Type", "Layer", "Value", "Thickness (mm)"])
    for table in dataset["pavement_tables"]:
        for layer in table["layers"]:
            ws_pavement.append(
                [
                    table["stretch_name"],
                    table["engineering_type"],
                    layer.get("layer"),
                    layer.get("value"),
                    layer.get("thickness_mm"),
                ]
            )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"road_project_{project_id}_dpr.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/projects/{project_id}/export/pdf")
def export_dpr_pdf(
    project_id: int,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    dataset = build_dpr_dataset(db, project_id)
    project = dataset["project"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"DPR Report - {project.name}", styles["Heading1"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Stretches", styles["Heading2"]))
    table_data = [["Stretch", "Category", "Engineering", "Start", "End", "Length (m)"]]
    for item in dataset["stretches"]:
        stretch = item["stretch"]
        table_data.append(
            [
                stretch.stretch_name,
                stretch.road_category,
                stretch.engineering_type,
                stretch.start_chainage,
                stretch.end_chainage,
                stretch.length_m,
            ]
        )
    stretch_table = Table(table_data, repeatRows=1)
    stretch_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    elements.append(stretch_table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Pavement Composition", styles["Heading2"]))
    pav_data = [["Stretch", "Engineering", "Layer", "Value", "Thickness (mm)"]]
    for table in dataset["pavement_tables"]:
        for layer in table["layers"]:
            pav_data.append(
                [
                    table["stretch_name"],
                    table["engineering_type"],
                    layer.get("layer"),
                    layer.get("value"),
                    layer.get("thickness_mm"),
                ]
            )
    pav_table = Table(pav_data, repeatRows=1)
    pav_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    elements.append(pav_table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"road_project_{project_id}_dpr.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
